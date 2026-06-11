import asyncio
import csv
import sys
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from user.models import Profile, User, UserTaskManager


BASE_DIR = Path(__file__).resolve().parents[3]
MODULES_DIR = BASE_DIR / "modules"

if str(MODULES_DIR) not in sys.path:
    sys.path.insert(0, str(MODULES_DIR))

from browser.browser_session import BrowserSession
from clients.multilogin_client import MultiloginApiError, MultiloginClient
from services.task_logger import TaskLogger


class Command(BaseCommand):
    help = (
        "Create Multilogin browser profiles and save ml_profile_id / ml_folder_id "
        "into the existing Django Profile model. Optionally run custom Playwright "
        "logic inside the same command after profile creation."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--profile-id",
            type=int,
            help="Existing local Profile ID. If passed, command updates this Profile.",
        )
        parser.add_argument(
            "--profile-email",
            help="Existing local Profile email. Alternative to --profile-id.",
        )

        parser.add_argument(
            "--user-id",
            type=int,
            help="Existing User ID with Multilogin credentials.",
        )
        parser.add_argument(
            "--user-email",
            help="Existing User email with Multilogin credentials. Alternative to --user-id.",
        )

        parser.add_argument(
            "--profile-name",
            help="Profile name. If not passed, account email will be used by default.",
        )
        parser.add_argument(
            "--account-email",
            help="Account email to save into Profile.email. Also used as default profile name.",
        )

        parser.add_argument(
            "--csv",
            type=Path,
            help="CSV file with profiles.",
        )
        parser.add_argument(
            "--xlsx",
            type=Path,
            help="Excel .xlsx file with profiles.",
        )

        parser.add_argument(
            "--folder-id",
            required=True,
            help="Default Multilogin folder ID. File column ml_folder_id or folder_id can override it.",
        )

        parser.add_argument(
            "--browser-type",
            default="mimic",
            choices=("mimic", "stealthfox"),
        )
        parser.add_argument(
            "--os-type",
            default="windows",
            choices=("windows", "macos", "linux"),
        )
        parser.add_argument(
            "--start-url",
            default="https://www.trustpilot.com/",
        )
        parser.add_argument(
            "--core-version",
            type=int,
            default=None,
        )
        parser.add_argument(
            "--auto-update-core",
            action="store_true",
            default=None,
        )
        parser.add_argument(
            "--no-auto-update-core",
            action="store_false",
            dest="auto_update_core",
        )
        parser.add_argument(
            "--timeout",
            type=int,
            default=60,
        )
        parser.add_argument(
            "--force",
            action="store_true",
            help="Allow replacing existing ml_profile_id on Profile.",
        )
        parser.add_argument(
            "--skip-existing",
            action="store_true",
            help="Skip Profile rows that already have ml_profile_id.",
        )

        parser.add_argument(
            "--run-playwright-after-create",
            action="store_true",
            help="After creating Multilogin profile, start it and run custom Playwright code.",
        )
        parser.add_argument(
            "--playwright-action",
            default="profile_created_playwright_hook",
            help="Action name for temporary UserTaskManager row.",
        )

        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Do not call Multilogin API and do not save changes.",
        )

    def handle(self, *args, **options):
        self._token_cache: dict[int, str] = {}

        rows = self._get_rows(options)

        created_count = 0
        skipped_count = 0
        failed_count = 0

        for index, row in enumerate(rows, start=1):
            try:
                result = self._process_row(index=index, row=row, options=options)
            except CommandError as exc:
                failed_count += 1
                self.stderr.write(self.style.ERROR(str(exc)))
                continue
            except Exception as exc:
                failed_count += 1
                self.stderr.write(
                    self.style.ERROR(
                        f"Row {index}: unexpected error: {type(exc).__name__}: {exc}"
                    )
                )
                continue

            if result == "created":
                created_count += 1
            elif result == "skipped":
                skipped_count += 1
            elif result == "failed":
                failed_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Done. Created: {created_count}. Skipped: {skipped_count}. Failed: {failed_count}."
            )
        )

    def _process_row(self, index: int, row: dict[str, str], options) -> str:
        profile = self._get_existing_profile(row)

        if profile:
            user = profile.user
            account_email = (
                self._first(row, "account_email", "email", "trustpilot_email")
                or profile.email
            )
            profile_name = (
                self._first(row, "profile_name", "name")
                or options.get("profile_name")
                or account_email
                or profile.profile_name
            )
        else:
            user = self._get_user(row, options)
            account_email = self._first(row, "account_email", "email", "trustpilot_email")
            profile_name = (
                self._first(row, "profile_name", "name")
                or options.get("profile_name")
                or account_email
            )

        if not user:
            raise CommandError(f"Row {index}: User was not found.")

        if not user.password:
            raise CommandError(
                f"Row {index}: User {user.email} does not have Multilogin password saved."
            )

        if not profile_name:
            raise CommandError(
                f"Row {index}: profile_name or account email is required."
            )

        folder_id = (
            self._first(row, "ml_folder_id", "folder_id")
            or options.get("folder_id")
        )

        if not folder_id:
            raise CommandError(
                f"Row {index}: ml_folder_id or --folder-id is required."
            )

        if profile and profile.ml_profile_id and not options["force"]:
            message = (
                f"Row {index}: Profile ID {profile.id} already has "
                f"ml_profile_id={profile.ml_profile_id}."
            )

            if options["skip_existing"]:
                self.stdout.write(self.style.WARNING(f"Skipped. {message}"))
                return "skipped"

            raise CommandError(message + " Use --force to replace it.")

        if not profile and account_email:
            existing_profile = Profile.objects.filter(email=account_email).first()

            if existing_profile and not options["force"]:
                message = (
                    f"Row {index}: Profile with email {account_email} already exists "
                    f"with ID {existing_profile.id}."
                )

                if options["skip_existing"]:
                    self.stdout.write(self.style.WARNING(f"Skipped. {message}"))
                    return "skipped"

                raise CommandError(message + " Use --force to update it.")

            if existing_profile and options["force"]:
                profile = existing_profile
                user = profile.user

        payload = MultiloginClient.build_profile_payload(
            name=profile_name,
            folder_id=folder_id,
            browser_type=options["browser_type"],
            os_type=options["os_type"],
            start_url=self._first(row, "start_url") or options["start_url"],
            core_version=options["core_version"],
            auto_update_core=options["auto_update_core"],
            notes=self._first(row, "notes") or f"Django Profile: {profile_name}",
            proxy=self._build_proxy(row),
        )

        if options["dry_run"]:
            self.stdout.write(
                self.style.WARNING(
                    f"DRY RUN row {index}: would create Multilogin profile "
                    f"'{profile_name}' for User {user.email} in folder {folder_id}."
                )
            )
            return "skipped"

        try:
            token = self._get_token(user=user, timeout=options["timeout"])
        except MultiloginApiError as exc:
            raise CommandError(
                f"Row {index}: failed to get Multilogin token for User {user.email}: {exc}"
            ) from exc

        try:
            result = MultiloginClient.create_profile_with_token(
                token=token,
                payload=payload,
                timeout=options["timeout"],
            )
        except MultiloginApiError as exc:
            raise CommandError(
                f"Row {index}: failed to create Multilogin profile '{profile_name}': {exc}"
            ) from exc

        ml_profile_id = result.profile_ids[0]

        account_password = self._first(
            row,
            "account_password",
            "password",
            "trustpilot_password",
        )
        fa2_secret = self._first(
            row,
            "fa2_secret",
            "2fa_secret",
            "two_fa_secret",
        )

        with transaction.atomic():
            if profile:
                update_fields = []

                profile.profile_name = profile_name
                update_fields.append("profile_name")

                profile.ml_profile_id = ml_profile_id
                update_fields.append("ml_profile_id")

                profile.ml_folder_id = folder_id
                update_fields.append("ml_folder_id")

                if account_email:
                    profile.email = account_email
                    update_fields.append("email")

                if account_password:
                    profile.password = account_password
                    update_fields.append("password")

                if fa2_secret:
                    profile.fa2_secret = fa2_secret
                    update_fields.append("fa2_secret")

                profile.save(update_fields=list(dict.fromkeys(update_fields)))
            else:
                profile = Profile.objects.create(
                    user=user,
                    profile_name=profile_name,
                    email=account_email or None,
                    password=account_password or None,
                    fa2_secret=fa2_secret or None,
                    ml_profile_id=ml_profile_id,
                    ml_folder_id=folder_id,
                    stage_status=Profile.StageStatus.NOT_STARTED,
                )

        self.stdout.write(
            self.style.SUCCESS(
                f"Row {index}: created Multilogin profile {ml_profile_id} "
                f"and saved it to Django Profile ID {profile.id}."
            )
        )

        if options["run_playwright_after_create"]:
            self._run_playwright_after_create(
                profile=profile,
                row=row,
                options=options,
            )

        return "created"

    def _run_playwright_after_create(
        self,
        *,
        profile: Profile,
        row: dict[str, str],
        options,
    ) -> None:
        """
        Starts the created Multilogin profile and runs custom Playwright code.

        This hook is inside the same command:
        create_multilogin_browser_profile.py

        After profile creation, run the command with:

            --run-playwright-after-create

        Then write your own parser logic in _custom_playwright_parser().
        """

        task = UserTaskManager.objects.create(
            profile=profile,
            action=options["playwright_action"],
            task_status=UserTaskManager.Status.IN_PROGRESS,
            started_at=timezone.now(),
        )

        try:
            asyncio.run(
                self._run_playwright_after_create_async(
                    profile=profile,
                    row=row,
                    task=task,
                )
            )
        except Exception as exc:
            task.task_status = UserTaskManager.Status.FAILED
            task.finished_at = timezone.now()
            task.comment = f"Playwright hook failed: {type(exc).__name__}: {exc}"
            task.save(
                update_fields=[
                    "task_status",
                    "finished_at",
                    "comment",
                    "updated_at",
                ]
            )

            raise CommandError(
                f"Playwright hook failed for Profile ID {profile.id}: "
                f"{type(exc).__name__}: {exc}"
            ) from exc

        task.task_status = UserTaskManager.Status.SUCCESS
        task.finished_at = timezone.now()
        task.save(update_fields=["task_status", "finished_at", "updated_at"])

    async def _run_playwright_after_create_async(
        self,
        *,
        profile: Profile,
        row: dict[str, str],
        task: UserTaskManager,
    ) -> None:
        logger = TaskLogger(task)

        browser_session = BrowserSession(
            task=task,
            multilogin_client=MultiloginClient(logger),
            logger=logger,
        )

        try:
            page = await browser_session.start()

            await self._custom_playwright_parser(
                page=page,
                profile=profile,
                row=row,
                task=task,
                logger=logger,
            )
        finally:
            await browser_session.stop()

    async def _custom_playwright_parser(
        self,
        *,
        page,
        profile: Profile,
        row: dict[str, str],
        task: UserTaskManager,
        logger: TaskLogger,
    ) -> None:
        """
        ПИШЕШЬ СВОЙ PLAYWRIGHT-КОД ЗДЕСЬ.

        На этом моменте:

        - Multilogin profile уже создан;
        - profile.ml_profile_id уже сохранен;
        - profile.ml_folder_id уже сохранен;
        - браузер Multilogin уже запущен;
        - Playwright уже подключен к этому браузеру;
        - page — это готовый Playwright Page.

        Пример:

            await page.goto(
                "https://www.trustpilot.com/",
                wait_until="domcontentloaded",
                timeout=60000,
            )

            title = await page.title()
            await logger.info(f"Page title: {title}")

            cards = page.locator(".some-card-selector")
            count = await cards.count()

            for index in range(count):
                text = await cards.nth(index).inner_text()
                await logger.info(text)

        Важно:
        здесь использовать только async Playwright API.
        """

        await logger.info(
            f"Playwright hook is ready for Profile ID {profile.id}. "
            "Add your custom parser code inside _custom_playwright_parser()."
        )




        await page.goto(
            "https://www.microsoft.com/en-us/microsoft-365/outlook/log-in",
            wait_until="domcontentloaded",
            timeout=60000,
        )
        sign_in = page.locator("#c-shellmenu_custom_outline_signin_bhvr100_right")
        await sign_in.wait_for(state="visible", timeout=30000)
        await sign_in.click()

        await page.wait_for_timeout(15000)

        await page.locator('input[name="loginfmt"]').fill("MiloNiland03@outlook.com")

        await page.locator("#idSIButton9").click()

        await page.wait_for_timeout(15000)

        await page.locator("#passwordEntry").fill("SLP9ktvZ1")

        await page.locator('button[type="submit"]').click()

        await page.wait_for_timeout(20000)

        title = await page.title()
        await logger.info(f"Page title: {title}")

        await page.goto(
            "https://www.trustpilot.com/",
            wait_until="domcontentloaded",
            timeout=60000,
        )

        await page.wait_for_timeout(15000)

        # Клик по Log in
        login_link = page.locator('a[data-navigation-login-desktop-link="true"]')
        await login_link.wait_for(state="visible", timeout=30000)
        await login_link.click()


        await page.wait_for_timeout(15000)
        # Клик по первому Continue with email
        continue_email = page.locator('button[data-reveal-email-flow-button="true"]')
        await continue_email.wait_for(state="visible", timeout=30000)
        await continue_email.click()


        await page.wait_for_timeout(15000)
        # Ввод email
        email_lookup_input = page.locator('input#email-lookup[data-email-input="true"]')
        await email_lookup_input.wait_for(state="visible", timeout=30000)
        await email_lookup_input.fill("your@email.com")


        await page.wait_for_timeout(15000)
        # Клик по второму Continue with email
        email_lookup_button = page.locator('button[data-email-lookup-button="true"]')
        await email_lookup_button.wait_for(state="visible", timeout=30000)
        await email_lookup_button.click()


        await page.wait_for_timeout(15000)

    def _get_rows(self, options) -> list[dict[str, str]]:
        csv_path = options.get("csv")
        xlsx_path = options.get("xlsx")

        if csv_path and xlsx_path:
            raise CommandError("Use only one option: --csv or --xlsx.")

        if csv_path:
            return self._get_rows_from_csv(csv_path)

        if xlsx_path:
            return self._get_rows_from_xlsx(xlsx_path)

        return [
            {
                "profile_id": str(options.get("profile_id") or ""),
                "profile_email": options.get("profile_email") or "",
                "user_id": str(options.get("user_id") or ""),
                "user_email": options.get("user_email") or "",
                "profile_name": options.get("profile_name") or "",
                "account_email": options.get("account_email") or "",
                "ml_folder_id": options.get("folder_id") or "",
                "start_url": options.get("start_url") or "",
            }
        ]

    def _get_rows_from_csv(self, csv_path: Path) -> list[dict[str, str]]:
        if not csv_path.exists():
            raise CommandError(f"CSV file was not found: {csv_path}")

        with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)

            if not reader.fieldnames:
                raise CommandError("CSV file is empty or has no headers.")

            rows = []

            for raw_row in reader:
                row = {
                    self._normalize_key(key): self._clean(value)
                    for key, value in raw_row.items()
                    if key
                }

                if any(row.values()):
                    rows.append(row)

            return rows

    def _get_rows_from_xlsx(self, xlsx_path: Path) -> list[dict[str, str]]:
        if not xlsx_path.exists():
            raise CommandError(f"Excel file was not found: {xlsx_path}")

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = workbook.active

        all_rows = list(sheet.iter_rows(values_only=True))

        if not all_rows:
            raise CommandError("Excel file is empty.")

        headers = [
            self._normalize_key(value)
            for value in all_rows[0]
            if value is not None and str(value).strip()
        ]

        if not headers:
            raise CommandError("Excel file does not contain headers in the first row.")

        rows = []

        for raw_row in all_rows[1:]:
            row = {}

            for index, header in enumerate(headers):
                value = raw_row[index] if index < len(raw_row) else ""
                row[header] = self._clean(value)

            if any(row.values()):
                rows.append(row)

        return rows

    def _get_existing_profile(self, row: dict[str, str]) -> Profile | None:
        profile_id = self._first(row, "profile_id", "local_profile_id", "id")
        profile_email = self._first(
            row,
            "profile_email",
            "email",
            "account_email",
            "trustpilot_email",
        )

        if profile_id:
            try:
                return Profile.objects.select_related("user").get(id=profile_id)
            except Profile.DoesNotExist as exc:
                raise CommandError(f"Profile with id={profile_id} was not found.") from exc

        if profile_email:
            return Profile.objects.select_related("user").filter(email=profile_email).first()

        return None

    def _get_user(self, row: dict[str, str], options) -> User:
        user_id = (
            self._first(row, "user_id", "multilogin_user_id")
            or options.get("user_id")
        )
        user_email = (
            self._first(row, "user_email", "multilogin_user_email")
            or options.get("user_email")
        )

        if user_id and user_email:
            raise CommandError("Use only one User selector: user_id or user_email.")

        try:
            if user_id:
                return User.objects.get(id=user_id)

            if user_email:
                return User.objects.get(email=user_email)
        except User.DoesNotExist as exc:
            raise CommandError("User was not found.") from exc

        raise CommandError(
            "Pass --user-id / --user-email, or provide user_id / user_email in the file."
        )

    def _get_token(self, user: User, timeout: int) -> str:
        if user.id in self._token_cache:
            return self._token_cache[user.id]

        token = MultiloginClient.get_token_by_credentials(
            email=user.email,
            password=user.password,
            timeout=timeout,
        )

        self._token_cache[user.id] = token
        return token

    def _build_proxy(self, row: dict[str, str]) -> dict[str, Any] | None:
        proxy_host = self._first(row, "proxy_host", "proxy_address", "host")

        if not proxy_host:
            return None

        proxy_port = self._first(row, "proxy_port", "port")

        if not proxy_port:
            raise CommandError("proxy_port is required when proxy_host is provided.")

        try:
            port = int(proxy_port)
        except ValueError as exc:
            raise CommandError("proxy_port must be an integer.") from exc

        proxy = {
            "type": self._first(row, "proxy_type") or "http",
            "host": proxy_host,
            "port": port,
            "save_traffic": False,
        }

        proxy_username = self._first(row, "proxy_username", "proxy_login", "username")
        proxy_password = self._first(row, "proxy_password")

        if proxy_username:
            proxy["username"] = proxy_username

        if proxy_password:
            proxy["password"] = proxy_password

        return proxy

    def _first(self, row: dict[str, str], *keys: str) -> str:
        for key in keys:
            value = row.get(key)

            if value is not None and str(value).strip():
                return str(value).strip()

        return ""

    def _clean(self, value) -> str:
        if value is None:
            return ""

        return str(value).strip()

    def _normalize_key(self, value) -> str:
        return (
            str(value)
            .strip()
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
        )