from django.contrib import admin, messages
from django import forms
from django.contrib.admin.helpers import ActionForm
from django.contrib.admin.widgets import AutocompleteSelect
from django.utils.html import format_html, mark_safe
from django.utils import timezone
from django.db.models import Sum, Count, Q
from django.template.response import TemplateResponse
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .models import Profile, UserTaskManager, User, UserTaskLog, ProfileDayProgress, Review, ReviewStatisticsProxy, ProfileSessionEvent


class UserTaskActionForm(ActionForm):
    custom_action_name = forms.CharField(
        required=False, 
        label="Action Name", 
        widget=forms.TextInput(attrs={'placeholder': 'Enter action...'})
    )


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'name',
        'email',
        'date_created',
    )
    list_filter = (
        'date_created',
    )
    search_fields = (
        'name',
        'email',
    )
    ordering = ('-date_created',)


@admin.register(Profile)
class ProfileAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'profile_name',
        'email',
        'colored_stage_status',
        'current_day',
        'current_stage',
        'last_completed_day',
        'next_action_at',
        'total_successful_days',
        'total_failed_days',
        'interactions_count',
        'reviews_count',
        'date_created',
    )
    search_fields = (
        'profile_name',
        'email',
        'ml_profile_id',
    )
    list_filter = (
        'stage_status',
        'current_day',
        'current_stage',
        'next_action_at',
        'user',
    )
    ordering = ('-date_created',)
    actions = [
        'pause_selected',
        'resume_selected',
        'retry_current_day',
        'move_to_next_day',
        'move_to_previous_day',
        'reset_to_day_1',
        'mark_as_completed',
    ]

    class Media:
        css = {
            'all': ('admin/custom.css',)
        }

    @admin.display(description="Status", ordering="stage_status")
    def colored_stage_status(self, obj):
        classes = {
            "not_started": "status-new",
            "pending": "status-pending",
            "in_progress": "status-in-progress",
            "success": "status-success",
            "failed": "status-failed",
            "retry_required": "status-warning",
            "paused": "status-on-hold",
            "completed": "status-ready",
        }
        css_class = classes.get(str(obj.stage_status).lower(), "status-default")
        return format_html(
            '<span class="status-badge {}">{}</span>',
            css_class,
            obj.get_stage_status_display().upper(),
        )

    @admin.action(description="Pause selected profiles")
    def pause_selected(self, request, queryset):
        updated = queryset.update(stage_status=Profile.StageStatus.PAUSED)
        self.message_user(request, f"{updated} profile(s) paused.", level=messages.SUCCESS)

    @admin.action(description="Resume selected profiles")
    def resume_selected(self, request, queryset):
        updated = queryset.update(stage_status=Profile.StageStatus.PENDING)
        self.message_user(request, f"{updated} profile(s) resumed.", level=messages.SUCCESS)

    @admin.action(description="Retry current day")
    def retry_current_day(self, request, queryset):
        updated = queryset.update(
            stage_status=Profile.StageStatus.PENDING,
            failed_at=None,
            failure_reason=None,
        )
        self.message_user(request, f"{updated} profile(s) set to retry current day.", level=messages.SUCCESS)

    @admin.action(description="Move to next day")
    def move_to_next_day(self, request, queryset):
        updated = 0
        for profile in queryset:
            if (profile.current_day or 1) >= 21:
                profile.stage_status = Profile.StageStatus.COMPLETED
                profile.completed_at = timezone.now()
            else:
                profile.current_day = (profile.current_day or 1) + 1
                profile.current_stage = f"day_{profile.current_day}"
                profile.stage_status = Profile.StageStatus.PENDING
            profile.save()
            updated += 1
        self.message_user(request, f"{updated} profile(s) moved to next day.", level=messages.SUCCESS)

    @admin.action(description="Move to previous day")
    def move_to_previous_day(self, request, queryset):
        updated = 0
        for profile in queryset:
            current = profile.current_day or 1
            if current > 1:
                profile.current_day = current - 1
                profile.current_stage = f"day_{profile.current_day}"
                profile.stage_status = Profile.StageStatus.PENDING
                profile.save()
                updated += 1
        self.message_user(request, f"{updated} profile(s) moved to previous day.", level=messages.SUCCESS)

    @admin.action(description="Reset to Day 1")
    def reset_to_day_1(self, request, queryset):
        updated = queryset.update(
            current_day=1,
            current_stage="day_1",
            stage_status=Profile.StageStatus.PENDING,
            last_completed_day=None,
            failed_at=None,
            failure_reason=None,
            completed_at=None,
            total_successful_days=0,
            total_failed_days=0,
        )
        self.message_user(request, f"{updated} profile(s) reset to Day 1.", level=messages.SUCCESS)

    @admin.action(description="Mark as completed")
    def mark_as_completed(self, request, queryset):
        updated = queryset.update(
            stage_status=Profile.StageStatus.COMPLETED,
            completed_at=timezone.now(),
        )
        self.message_user(request, f"{updated} profile(s) marked as completed.", level=messages.SUCCESS)


@admin.register(UserTaskManager)
class UserTaskManagerAdmin(admin.ModelAdmin):
    action_form = UserTaskActionForm
    autocomplete_fields = ('profile',)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'profile':
            kwargs['widget'] = AutocompleteSelect(
                db_field, self.admin_site, attrs={'data-width': '45em'}
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    list_display = (
        'id',
        'action',
        'colored_status',
        'profile',
        'started_at',
        'finished_at',
        'duration',
        'interactions_count',
        'comment',
        'execute_at',
        'updated_at',
        'company_queries',
        'company_queries_count_from',
        'company_queries_count_to',
        'interactions_count_from',
        'interactions_count_to',
        'limit_on_reviews_duration',
        'pause_multiplier',
        'prompt_parameters',
        'prompts',
        'restricted_company_queries',
        'review_length',
        'styled_target_company',
        'profile_id',
        'maximum_longer_reviews',
        'maximum_review_pages',
    )
    list_filter = (
        'task_status',
        'action',
        'execute_at',
        'profile',
        'updated_at',
        'started_at',
        'finished_at',
        'target_company',
    )
    search_fields = (
        'profile__profile_name',
        'action',
        'target_company',
        'comment',
    )
    readonly_fields = (
        'duration',
        'started_at',
        'finished_at',
        'execute_at',
        'updated_at',
    )
    ordering = ('-execute_at',)

    class Media:
        css = {
            'all': ('admin/custom.css',)
        }

    @admin.display(description="Status", ordering="task_status")
    def colored_status(self, obj):
        classes = {
            "on_hold": "status-on-hold",
            "pending": "status-pending",
            "in_progress": "status-in-progress",
            "success": "status-success",
            "failed": "status-failed",
        }
        css_class = classes.get(str(obj.task_status).lower(), "status-default")
        return format_html(
            '<span class="status-badge {}">{}</span>',
            css_class,
            obj.get_task_status_display().upper(),
        )

    @admin.display(description="Target Company", ordering="target_company")
    def styled_target_company(self, obj):
        if not obj.target_company:
            return "-"
        return format_html(
            '<span class="column-highlight">{}</span>',
            obj.target_company
        )
    
    actions = [
        "duplicate_selected", 
        "mark_pending",
        "mark_in_progress",
        "mark_success",
        "mark_failed",
        "mark_on_hold",
        "clear_comment",
        "set_custom_action",
    ]

    @admin.action(description="Duplicate selected User tasks")
    def duplicate_selected(self, request, queryset):
        created = 0

        for obj in queryset:
            obj.pk = None
            obj.save()
            created += 1

        self.message_user(
            request,
            f"{created} item(s) duplicated.",
            level=messages.SUCCESS,
        )

    @admin.action(description="Set status = Pending")
    def mark_pending(self, request, queryset):
        updated = queryset.update(task_status="pending")
        self.message_user(request, f"{updated} row(s) updated.", level=messages.SUCCESS)

    @admin.action(description="Set status = In Progress")
    def mark_in_progress(self, request, queryset):
        updated = queryset.update(task_status="in_progress")
        self.message_user(request, f"{updated} row(s) updated.", level=messages.SUCCESS)

    @admin.action(description="Set status = Success")
    def mark_success(self, request, queryset):
        updated = queryset.update(task_status="success")
        self.message_user(request, f"{updated} row(s) updated.", level=messages.SUCCESS)

    @admin.action(description="Set status = Failed")
    def mark_failed(self, request, queryset):
        updated = queryset.update(task_status="failed")
        self.message_user(request, f"{updated} row(s) updated.", level=messages.SUCCESS)

    @admin.action(description="Set status = On Hold")
    def mark_on_hold(self, request, queryset):
        updated = queryset.update(task_status="on_hold")
        self.message_user(request, f"{updated} row(s) updated.", level=messages.SUCCESS)

    @admin.action(description="Clear comment")
    def clear_comment(self, request, queryset):
        updated = queryset.update(comment="")
        self.message_user(request, f"{updated} row(s) updated.", level=messages.SUCCESS)

    @admin.action(description="Set custom action name (from field)")
    def set_custom_action(self, request, queryset):
        new_action = request.POST.get('custom_action_name')
        if not new_action:
            self.message_user(request, "Error: You must type an action name in the field.", level=messages.ERROR)
            return
        updated = queryset.update(action=new_action)
        self.message_user(request, f"{updated} row(s) updated to '{new_action}'.", level=messages.SUCCESS)


@admin.register(UserTaskLog)
class UserTaskLogAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'colored_level',
        'message',
        'comment',
        'created_at',
        'task',
        'task_id',
    )
    list_filter = (
        'level',
        'created_at',
        'task',
    )
    search_fields = (
        'task__action',
        'message',
        'comment',
    )
    ordering = ('-created_at',)
    readonly_fields = ('created_at',)

    class Media:
        css = {
            'all': ('admin/custom.css',)
        }

    @admin.display(description="Level", ordering="level")
    def colored_level(self, obj):
        classes = {
            "info": "level-info",
            "warning": "level-warning",
            "error": "level-error",
            "debug": "level-debug",
        }
        # Use lowercase for lookup to be safe
        css_class = classes.get(str(obj.level).lower(), "status-default")
        return format_html(
            '<span class="status-badge {}">{}</span>',
            css_class,
            obj.get_level_display().upper(),
        )

    @admin.display(description="Message")
    def message_short(self, obj):
        return obj.message[:100] + "..." if len(obj.message) > 100 else obj.message


@admin.register(ProfileDayProgress)
class ProfileDayProgressAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'profile',
        'day_number',
        'scenario_name',
        'status',
        'started_at',
        'finished_at',
        'duration_seconds',
        'retry_count',
        'interactions_count',
        'reviews_count',
        'error_message',
    )
    list_filter = (
        'day_number',
        'scenario_name',
        'status',
        'started_at',
        'finished_at',
    )
    search_fields = (
        'profile__profile_name',
        'scenario_name',
        'error_message',
    )
    ordering = ('-created_at',)
    readonly_fields = ('created_at', 'updated_at')


@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    autocomplete_fields = ('profile',)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'profile':
            kwargs['widget'] = AutocompleteSelect(
                db_field, self.admin_site, attrs={'data-width': '45em'}
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    list_display = (
        'id',
        'account_email',
        'trustpilot_username',
        'company',
        'rating_stars',
        'status_badge',
        'review_title',
        'review_text_short',
        'review_link_clickable',
        'created_at',
    )
    list_filter = (
        'status',
        'rating',
        'company',
        'created_at',
        'profile',
    )
    search_fields = (
        'profile__profile_name',
        'profile__email',
        'trustpilot_username',
        'company',
        'review_title',
        'review_text',
    )
    ordering = ('-created_at',)
    readonly_fields = ()
    save_as = True
    fieldsets = (
        ('Account', {
            'fields': ('profile', 'task', 'trustpilot_username'),
        }),
        ('Review content', {
            'fields': ('company', 'rating', 'review_title', 'review_text', 'review_link'),
        }),
        ('Status', {
            'fields': ('status',),
            'description': (
                'Set to Successful only when the review was actually published on Trustpilot. '
                'Automation sets this automatically. Manual entries must select a value explicitly.'
            ),
        }),
        ('Metadata', {
            'fields': ('created_at',),
        }),
    )

    actions = ['export_to_excel', 'mark_successful', 'mark_unsuccessful', 'clear_status']

    @admin.action(description='Export selected reviews to Excel (.xlsx)')
    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reviews'

        headers = [
            'ID', 'Email', 'Trustpilot Username', 'Company', 'Rating',
            'Status', 'Review Title', 'Review Text', 'Link', 'Created At',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = 'A2'

        for review in queryset.select_related('profile'):
            ws.append([
                review.id,
                review.profile.email if review.profile else '',
                review.trustpilot_username or '',
                review.company or '',
                review.rating,
                review.status or '',
                review.review_title or '',
                review.review_text or '',
                review.review_link or '',
                review.created_at.strftime('%Y-%m-%d %H:%M:%S') if review.created_at else '',
            ])

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[letter].width = min(width + 2, 80)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="reviews_export.xlsx"'
        wb.save(response)
        return response

    @admin.action(description='Mark selected reviews as Successful')
    def mark_successful(self, request, queryset):
        updated = queryset.update(status=Review.Status.SUCCESSFUL)
        self.message_user(request, f'{updated} review(s) marked as Successful.')

    @admin.action(description='Mark selected reviews as Unsuccessful')
    def mark_unsuccessful(self, request, queryset):
        updated = queryset.update(status=Review.Status.UNSUCCESSFUL)
        self.message_user(request, f'{updated} review(s) marked as Unsuccessful.')

    @admin.action(description='Clear status (set to unknown)')
    def clear_status(self, request, queryset):
        updated = queryset.update(status=None)
        self.message_user(request, f'{updated} review(s) status cleared.')

    class Media:
        css = {'all': ('admin/custom.css',)}

    @admin.display(description='Email', ordering='profile__email')
    def account_email(self, obj):
        return obj.profile.email or '—'

    @admin.display(description='Rating', ordering='rating')
    def rating_stars(self, obj):
        if obj.rating is None:
            return '—'
        stars = '★' * obj.rating + '☆' * (5 - obj.rating)
        color = '#059669' if obj.rating >= 4 else ('#d97706' if obj.rating == 3 else '#dc2626')
        return format_html('<span style="color:{};font-size:14px;">{}</span>', color, stars)

    @admin.display(description='Status', ordering='status')
    def status_badge(self, obj):
        if obj.status == Review.Status.SUCCESSFUL:
            return mark_safe('<span class="status-badge status-success">SUCCESSFUL</span>')
        if obj.status == Review.Status.UNSUCCESSFUL:
            return mark_safe('<span class="status-badge status-failed">UNSUCCESSFUL</span>')
        return mark_safe('<span class="status-badge status-pending">—</span>')

    @admin.display(description='Review text')
    def review_text_short(self, obj):
        if not obj.review_text:
            return '—'
        text = obj.review_text[:120]
        return f'{text}…' if len(obj.review_text) > 120 else text

    @admin.display(description='Link')
    def review_link_clickable(self, obj):
        if not obj.review_link:
            return '—'
        return format_html('<a href="{}" target="_blank">Open</a>', obj.review_link)


@admin.register(ReviewStatisticsProxy)
class ReviewStatisticsAdmin(admin.ModelAdmin):
    """Custom statistics page. Uses a proxy model so it appears naturally in the admin sidebar."""

    class Media:
        css = {'all': ('admin/custom.css',)}

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_urls(self):
        from django.urls import path
        info = self.model._meta.app_label, self.model._meta.model_name
        return [
            path(
                '',
                self.admin_site.admin_view(self.statistics_view),
                name='%s_%s_changelist' % info,
            ),
        ]

    def statistics_view(self, request):
        from django.urls import reverse
        review_list_base = reverse('admin:user_review_changelist')

        total_reviews = Review.objects.count()
        successful_reviews = Review.objects.filter(status=Review.Status.SUCCESSFUL).count()
        unsuccessful_reviews = Review.objects.filter(status=Review.Status.UNSUCCESSFUL).count()

        # Per-profile breakdown with per-status counts in a single query.
        profiles = (
            Profile.objects
            .annotate(
                review_count=Count('reviews'),
                successful_count=Count('reviews', filter=Q(reviews__status=Review.Status.SUCCESSFUL)),
                unsuccessful_count=Count('reviews', filter=Q(reviews__status=Review.Status.UNSUCCESSFUL)),
            )
            .filter(review_count__gt=0)
            .select_related('user')
            .order_by('-review_count')
        )
        profiles_with_links = [
            {
                'profile': p,
                'review_count': p.review_count,
                'successful_count': p.successful_count,
                'unsuccessful_count': p.unsuccessful_count,
                'edit_url': f'{review_list_base}?profile__id__exact={p.pk}',
            }
            for p in profiles
        ]

        context = {
            **self.admin_site.each_context(request),
            'title': 'Reviews Statistics',
            'total_reviews': total_reviews,
            'successful_reviews': successful_reviews,
            'unsuccessful_reviews': unsuccessful_reviews,
            'profiles': profiles_with_links,
            'add_review_url': reverse('admin:user_review_add'),
            'opts': self.model._meta,
        }
        return TemplateResponse(request, 'admin/review_statistics.html', context)


@admin.register(ProfileSessionEvent)
class ProfileSessionEventAdmin(admin.ModelAdmin):
    autocomplete_fields = ('profile',)
    list_display = (
        'id',
        'profile_link',
        'warming_day',
        'date',
        'session_status_badge',
        'session_dropped_at',
        'error_reason_short',
        'notes',
        'created_at',
    )
    list_filter = (
        'session_status',
        'warming_day',
        'date',
        'profile',
    )
    search_fields = (
        'profile__profile_name',
        'profile__email',
        'error_reason',
        'notes',
    )
    ordering = ('-created_at',)
    readonly_fields = ('created_at', 'updated_at')
    date_hierarchy = 'date'

    class Media:
        css = {'all': ('admin/custom.css',)}

    @admin.display(description='Profile', ordering='profile__profile_name')
    def profile_link(self, obj):
        from django.urls import reverse as _reverse
        url = _reverse('admin:user_profile_change', args=[obj.profile_id])
        return format_html('<a href="{}">{}</a>', url, obj.profile.profile_name)

    @admin.display(description='Status', ordering='session_status')
    def session_status_badge(self, obj):
        classes = {
            'alive': 'status-success',
            'expired': 'status-warning',
            'relogin_required': 'status-failed',
        }
        css = classes.get(obj.session_status, 'status-default')
        return format_html(
            '<span class="status-badge {}">{}</span>',
            css,
            obj.get_session_status_display().upper(),
        )

    @admin.display(description='Error reason')
    def error_reason_short(self, obj):
        if not obj.error_reason:
            return '—'
        return obj.error_reason[:100] + '…' if len(obj.error_reason) > 100 else obj.error_reason